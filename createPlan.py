# createPlan
import openpyxl.styles
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill

class Player:
    def __init__(self, name, days):
        self.name = name
        self.days = days

def create_dataframe(players, size, start):
    players.sort(key=lambda x: x.days, reverse=True)
    matrix = np.empty((size, 7), dtype=object)
    check_matrix = np.empty((size, 7), dtype=object)
    free_space = np.empty(size)

    #create first fill 
    for i in range(size):
        days_player = players[i].days
        if(days_player == 7):
            start = i
        name = players[i].name
        for j in range(7):
            matrix[i][j] = name
            free_space[i] = 7 - days_player

    #fill with rest
    for x in range (size, len(players)):
        days_player = players[x].days
        name = players[x].name
        free_days = np.empty(7, dtype=object)
        for a in range (start, size):
            for b in range(7):
                if(days_player > 0 and check_matrix[a][b] is None and free_space[a] > 0 and free_days[b] is None):
                    matrix[a][b] = name
                    check_matrix[a][b] = 1 
                    free_space[a] -= 1
                    days_player -= 1
                    free_days[b] = 1


    # Print the matrix (for debugging purposes)
    '''
    i = 0
    for row in matrix:
        print(f"{row} {free_space[i]}")
        i += 1
    '''
    columns = [f'Day {i+1}' for i in range(7)]

    df = pd.DataFrame(matrix, columns=columns)
    color_df(df, start, players, size, "roster_plan.xlsx")
    return df

p = Player  # Assigning Player class to p for shorthand use
players = [
    p("P1", 7), p("P2", 7), p("P3", 7),
    p("P4", 7), p("P5", 7), p("P6", 7),
    p("P7", 7), p("P8", 7), p("P9", 7),
    p("P10", 6), p("P11", 6), p("P12", 5),
    p("P13", 4), p("P14", 3), p("P15", 2),
    p("P16", 2), p("P17", 2), p("P18", 2),
    p("P19", 2), p("P20", 1), p("P0", 7)
]

def color_df(df, start, players, size, path):
    df.to_excel(path)
    wb = openpyxl.load_workbook(path)
    ws = wb['Sheet1']

    # fill all that play 7 (blue)
    fill_pattern_blue = PatternFill(patternType='solid', fgColor = 'FF51A2E8') #blue
    for row in range(2, start + 2): 
        for col in range(2, 9):  
            cell = ws.cell(row=row, column=col)
            cell.fill = fill_pattern_blue
            
    # fill the row ones alternately
    fill_pattern_a = PatternFill(patternType='solid', fgColor = 'FFE89F51') #yellow
    fill_pattern_b = PatternFill(patternType='solid', fgColor = 'FFB9EA8C') #green
    for row in range(start+2, size + 2):  
        for col in range(2, 9):  
            cell = ws.cell(row=row, column=col)
            if(row%2 == 0): cell.fill = fill_pattern_a
            else: cell.fill = fill_pattern_b

    fill_patterns = [
    PatternFill(patternType='solid', fgColor='FFFF5733'),  # Red
    PatternFill(patternType='solid', fgColor='FF9B59B6'),  # Purple
    PatternFill(patternType='solid', fgColor='FF85C1E9'),  # Light Blue
    PatternFill(patternType='solid', fgColor='FF008080'),  # Teal
    PatternFill(patternType='solid', fgColor='FFFFCC55'),  # Yellow
    PatternFill(patternType='solid', fgColor='FF1ABC9C'),  # Dark Blue
    PatternFill(patternType='solid', fgColor='FFDFFF00'),  # Bright Yellow
    PatternFill(patternType='solid', fgColor='FF8E44AD'),  # Violet
    PatternFill(patternType='solid', fgColor='FF2980B9'),  # Deep Blue
    PatternFill(patternType='solid', fgColor='FF27AE60')   # Emerald Green
]

    # Iterate over players and assign each player a different color
    for idx, player in enumerate(players[size:]):
        # Use modulo to cycle through the fill_patterns if players exceed the number of patterns
        fill_pattern = fill_patterns[idx % len(fill_patterns)]

        # Call your function to apply the color
        color_name(player.name, fill_pattern, ws, start, size)


    wb.save(path)
    #print("done color df")

def color_name(name, fill, worksheet, start, size):
    for row in range(start+3, size + 2):  
        for col in range(2, 9):
            cell = worksheet.cell(row=row, column=col)
            if cell.value == name:
                cell.fill = fill

def create_excel_plan(roster_entries, size, file_path="roster_plan.xlsx"):
    players = []
    start = 0
    for entry in roster_entries:
        name_entry, days_entry, _ = entry  # Ignore the delete button
        # Changed from cget("text") to get() since these are CTkEntry widgets
        name = name_entry.get()
        days_str = days_entry.get()
        
        # Validate the days input
        if not days_str.isdigit():
            continue
            
        days = int(days_str)
        
        # Skip invalid entries
        if days < 1 or days > 7:
            continue
            
        new_player = Player(name, days)
        if days == 7: 
            start += 1
        players.append(new_player)
    
    # Check if we have enough players
    if len(players) < size:
        raise ValueError(f"Not enough valid players. Need {size}, but only have {len(players)}.")
    
    df = create_dataframe(players, size, start)
    color_df(df, start, players, size, file_path)

    #print(f"Excel file created successfully at {file_path}")
    

#create_dataframe(players, 15, "roster_plan.xlsx")